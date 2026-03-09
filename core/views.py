import logging
import os
from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.db import transaction
from django.core.files.storage import default_storage
from django.conf import settings
from openpyxl import Workbook
import openpyxl

from .models import Store, ESLTag, Gateway, TagHardware, Product
from .mqtt_client import mqtt_service
from .services import BulkMapProcessor, process_modisoft_file_logic
from .middleware import InputSanitizationMiddleware

"""
SAIS CORE VIEWS: REQUEST HANDLING & BUSINESS LOGIC
--------------------------------------------------
In Django, a 'View' is a Python function that takes a Web Request
and returns a Web Response (HTML, JSON, File, etc.).

Key Responsibilities:
1. STORE SELECTION: Managing which physical store the user is currently "working in".
2. DATA IMPORTS: Processing Excel files for products and tags.
3. BULK OPERATIONS: Handling barcode scanner data for tag pairing.
4. GATEWAY CONFIG: Pushing network settings to hardware via MQTT.

EDUCATIONAL: Decorators (like @login_required) are used to wrap functions
with extra logic (like security checks) before the main code runs.
"""

logger = logging.getLogger(__name__)

# --- DECORATORS (Pre-processing Logic) ---

def store_required(view_func):
    """
    CUSTOM DECORATOR
    ----------------
    Ensures that a user has actually selected a Store from the dropdown
    before they can access specific management pages.
    """
    @wraps(view_func)
    @login_required # First, ensure they are logged in
    def _wrapped_view(request, *args, **kwargs):
        # Check if 'active_store' was set by our Middleware
        if not hasattr(request, 'active_store') or request.active_store is None:
            messages.warning(request, "Please select a store first.")
            return redirect('select_store')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# --- STORE SELECTION (The Multi-Tenant Gateway) ---

@login_required
def select_store(request):
    """
    STORE PICKER PAGE
    -----------------
    Displays a list of stores the user is authorized to manage.
    This is the first thing a user sees after logging in.
    """
    try:
        if request.user.is_superuser:
            # Admins can see every active store in the system
            user_stores = Store.objects.filter(is_active=True).order_by('name')
            user_company = None
        else:
            user_company = getattr(request.user, 'company', None)
            if not user_company:
                return render(request, 'admin/core/no_access.html', {'reason': "User account not linked to any company."})

            if request.user.role == 'owner':
                # Owners see all stores within their company
                user_stores = Store.objects.filter(company=user_company, is_active=True).order_by('name')
            else:
                # Managers/Staff only see stores they are explicitly assigned to
                user_stores = request.user.managed_stores.filter(is_active=True).order_by('name')

        if not user_stores.exists():
            return render(request, 'admin/core/no_access.html', {'reason': 'No stores assigned to your account.'})

        # SHORTCUT: If they only have 1 store, just select it and move to the dashboard
        if user_stores.count() == 1:
            request.session['active_store_id'] = user_stores.first().id
            return redirect('admin:index')

        return render(request, 'admin/select_store.html', {'stores': user_stores, 'user_company': user_company})
    except Exception as e:
        logger.exception("Error in select_store view")
        messages.error(request, "An unexpected error occurred while loading stores.")
        return redirect('admin:index')

@login_required
def set_active_store(request, store_id):
    """
    ACTION: SWITCH STORE
    --------------------
    Updates the 'active_store_id' in the user's Session cookie.
    """
    try:
        # Security: Verify they actually have permission for the store they are trying to select
        if request.user.is_superuser:
            store = get_object_or_404(Store, id=store_id, is_active=True)
        else:
            user_company = getattr(request.user, 'company', None)
            if request.user.role == 'owner':
                store = get_object_or_404(Store, id=store_id, company=user_company, is_active=True)
            else:
                store = get_object_or_404(request.user.managed_stores.filter(is_active=True), id=store_id)

        # Write the ID to the session. The 'StoreContextMiddleware' will read
        # this on the next page load.
        request.session['active_store_id'] = store.id
        return redirect('admin:index')
    except Exception as e:
        logger.exception(f"Error setting active store {store_id}")
        messages.error(request, "Could not select the store. Please try again.")
        return redirect('select_store')

# --- DATA IMPORT VIEWS ---

@login_required
def download_tag_template(request):
    """
    ACTION: GET EXCEL TEMPLATE
    --------------------------
    Generates a sample Excel file on-the-fly using 'openpyxl' and
    sends it to the user's browser.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(['tag_mac', 'gateway_mac', 'model_name'])
        ws.append(['BE:01:02:03:04:05', 'FF:EE:DD:CC:BB:AA', 'Mi 05'])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=esl_tag_template.xlsx'
        wb.save(response)
        return response
    except Exception as e:
        logger.exception("Error downloading tag template")
        messages.error(request, "Failed to generate template.")
        return redirect('admin:core_esltag_changelist')

@login_required
def preview_tag_import(request):
    """
    ACTION: PREVIEW TAG UPLOAD
    --------------------------
    Parses a user-uploaded Excel file, identifies new vs. existing tags,
    and shows a preview table before committing any changes to the DB.
    """
    if request.method != 'POST' or not request.FILES.get('file'):
        return redirect('admin:core_esltag_changelist')
    
    active_store = getattr(request, 'active_store', None)
    if not active_store:
        messages.error(request, "Select a store first.")
        return redirect('admin:core_esltag_changelist')

    try:
        # Load the uploaded Excel file from memory
        wb = openpyxl.load_workbook(request.FILES['file'], data_only=True, read_only=True)
        sheet = wb.active

        summary = {'added': 0, 'updated': 0, 'rejected': 0, 'unchanged': 0}
        results = []

        # Iterate through rows starting from row 2 (skip headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row[:3]): continue

            # 1. Clean data: Remove spaces/special chars from MAC address
            sanitized_id = InputSanitizationMiddleware.sanitize_tag_id(row[0])

            # 2. Lookups: Find the physical spec and the gateway
            spec = TagHardware.objects.filter(model_number=str(row[2] or "").strip()).first()
            gateway = Gateway.objects.filter(gateway_mac__iexact=str(row[1] or ""), store=active_store).first()

            # 3. Validation: Reject if mandatory data is missing
            if not sanitized_id or not spec or not gateway:
                summary['rejected'] += 1
                results.append({'mac': str(row[0]), 'status': 'rejected', 'message': 'Invalid ID, Model, or Gateway.'})
                continue

            # 4. Save: Get or Create the record in the DB
            tag, created = ESLTag.objects.get_or_create(
                tag_mac=sanitized_id,
                store=active_store,
                defaults={'gateway': gateway, 'hardware_spec': spec, 'updated_by': request.user}
            )

            if created:
                summary['added'] += 1
                status, msg = 'added', "New tag registered."
            elif tag.gateway != gateway or tag.hardware_spec != spec:
                # Update existing record if something changed
                tag.gateway, tag.hardware_spec, tag.updated_by = gateway, spec, request.user
                tag.save()
                summary['updated'] += 1
                status, msg = 'updated', "Updated metadata."
            else:
                summary['unchanged'] += 1
                status, msg = 'unchanged', "No changes."

            results.append({'mac': sanitized_id, 'status': status, 'message': msg})

        # Return the 'Import Preview' HTML page
        return render(request, 'admin/core/esltag/import_preview.html', {
            'summary': summary,
            'results': results,
            'opts': ESLTag._meta
        })
    except Exception as e:
        logger.exception("Error processing tag import")
        messages.error(request, "An error occurred during import processing.")
        return redirect('admin:core_esltag_changelist')

@login_required
def preview_product_import(request):
    """
    ACTION: MODISOFT EXCEL IMPORT
    -----------------------------
    A multi-step import process for retail products.
    Step 1: Upload file.
    Step 2: Preview changes (prices going up/down).
    Step 3: Click 'Confirm' to commit to Database.
    """
    active_store = getattr(request, 'active_store', None)
    if not active_store:
        messages.error(request, "Select a store first.")
        return redirect('admin:core_product_changelist')

    try:
        if request.method == "POST":
            # PHASE 2: CONFIRMATION
            if "confirm_save" in request.POST:
                rel_filename = request.POST.get("temp_filename")
                if not rel_filename:
                    messages.error(request, "No file provided.")
                    return redirect('admin:core_product_changelist')

                # SECURITY: Verify the file path is safe and hasn't been tampered with
                normalized_name = os.path.normpath(rel_filename)
                if normalized_name.startswith('..') or not normalized_name.startswith('tmp' + os.sep):
                    logger.warning(f"Security: Blocked suspicious import path: {rel_filename}")
                    messages.error(request, "Invalid file.")
                    return redirect('admin:core_product_changelist')

                temp_path = os.path.join(settings.MEDIA_ROOT, normalized_name)

                # Call the 'Service' layer to perform the actual DB writes
                results, error = process_modisoft_file_logic(temp_path, active_store, request.user, commit=True)

                if not error:
                    # Clean up the temporary file
                    if os.path.exists(temp_path): os.remove(temp_path)
                    messages.success(request, f"Imported {len(results['new'])} new, updated {len(results['update'])} products.")
                    return redirect('admin:core_product_changelist')

                messages.error(request, error)

            # PHASE 1: UPLOAD & PREVIEW
            elif request.FILES.get("import_file"):
                myfile = request.FILES["import_file"]
                # Save file to a temporary location for review
                temp_filename = default_storage.save(os.path.join('tmp', myfile.name), myfile)
                temp_path = os.path.join(settings.MEDIA_ROOT, temp_filename)

                # Parse without committing to DB (commit=False)
                results, error = process_modisoft_file_logic(temp_path, active_store, request.user, commit=False)

                if error:
                    messages.error(request, error)
                    return redirect('admin:core_product_changelist')

                return render(request, "admin/core/product/import_preview.html", {
                    "results": results,
                    "temp_filename": temp_filename,
                    "store": active_store
                })
    except Exception as e:
        logger.exception("Error in product import view")
        messages.error(request, "An unexpected error occurred during product import.")
        return redirect('admin:core_product_changelist')

    # Default state: Show the upload form
    return render(request, "admin/core/product/import_upload.html", {"store": active_store})

@login_required
def bulk_map_tags_view(request):
    """
    ACTION: SCANNER LINKAGE
    ----------------------
    Handles raw data from a barcode scanner.
    A typical scan pattern is: [SKU] -> [TAG ID] -> [SKU] -> [TAG ID].
    """
    opts = ESLTag._meta
    context = {'opts': opts, 'app_label': opts.app_label, 'title': "Bulk Product-Tag Mapping"}

    try:
        if request.method == "POST":
            # ACTION: COMMIT MAPPINGS
            if 'confirm_mapping' in request.POST:
                proposed_data = request.session.get('pending_bulk_maps', [])
                # Use a TRANSACTION to ensure all updates happen together (all or nothing)
                with transaction.atomic():
                    for item in proposed_data:
                        # Update the database
                        ESLTag.objects.filter(id=item['tag_id']).update(
                            paired_product_id=item['product_id'],
                            updated_by=request.user
                        )
                        # Trigger an immediate hardware update for this tag
                        from .tasks import update_tag_image_task
                        update_tag_image_task.delay(item['tag_id'])

                messages.success(request, f"Successfully mapped {len(proposed_data)} tags.")
                if 'pending_bulk_maps' in request.session: del request.session['pending_bulk_maps']
                return redirect("admin:core_esltag_changelist")

            # ACTION: PROCESS RAW SCANS
            import_file = request.FILES.get('import_file')
            if not import_file: return redirect(request.path)

            raw_text = import_file.read().decode('utf-8')
            active_store = getattr(request, 'active_store', None)

            # Use 'BulkMapProcessor' (logic defined in services.py) to match SKUs to Tags
            processor = BulkMapProcessor(raw_text, active_store, request.user)
            proposed, rejections = processor.process()

            # Store the proposed mapping in the user's session for the next step
            request.session['pending_bulk_maps'] = proposed

            context.update({'proposed': proposed, 'rejections': rejections, 'stage': 'preview'})
            return render(request, 'admin/core/esltag/bulk_map_preview.html', context)

    except Exception as e:
        logger.exception("Error in bulk mapping view")
        messages.error(request, "An error occurred during bulk mapping.")
        return redirect('admin:core_esltag_changelist')

    return render(request, 'admin/core/esltag/bulk_map_upload.html', context)

@login_required
def configure_gateway_view(request, gateway_id):
    """
    ACTION: REMOTE GATEWAY SETUP
    ---------------------------
    A view specifically for pushing new IP/Server settings to a physical
    gateway via an MQTT command.
    """
    # Security: Only superusers can change hardware network config
    if not request.user.is_superuser:
        raise PermissionDenied

    gateway = get_object_or_404(Gateway, pk=gateway_id)
    opts = Gateway._meta

    if request.method == "POST":
        # Collect parameters from the form
        alias = request.POST.get('alias')
        server = request.POST.get('server') or "192.168.1.92:9081"
        username = request.POST.get('username') or "test"
        password = request.POST.get('password') or "123456"
        encrypt = request.POST.get('encrypt') == 'on'
        auto_ip = request.POST.get('auto_ip') == 'on'
        local_ip = request.POST.get('local_ip', '')
        netmask = request.POST.get('netmask', '')
        network_gateway = request.POST.get('network_gateway', '')

        # Basic validation: Ensure Server is 'IP:Port'
        import re
        ip_port_pattern = r'^(\d{1,3}\.){3}\d{1,3}:\d{1,5}$'
        if not re.match(ip_port_pattern, server):
            messages.error(request, f"Invalid Server format: {server}. Expected IP:Port (e.g., 192.168.1.92:9081)")
            return redirect(request.path)

        try:
            heartbeat = int(request.POST.get('heartbeat', 300))
        except ValueError:
            heartbeat = 300

        server_ip, server_port = server.split(':')

        # CALL MQTT: Send the physical command to the hardware
        success = mqtt_service.publish_config(
            gateway.estation_id,
            alias,
            server,
            encrypt,
            heartbeat,
            auto_ip=auto_ip,
            local_ip=local_ip,
            subnet=netmask,
            gateway=network_gateway,
            username=username,
            password=password
        )

        if success:
            messages.success(request, f"Configuration push for {gateway} initiated.")
            # Update local record to match what we just sent
            gateway.alias = alias
            gateway.app_server_ip = server_ip
            gateway.app_server_port = int(server_port)
            gateway.username = username
            gateway.password = password
            gateway.heartbeat_interval = heartbeat
            gateway.is_encrypt_enabled = encrypt
            gateway.is_auto_ip = auto_ip
            gateway.local_ip = local_ip
            gateway.netmask = netmask
            gateway.network_gateway = network_gateway
            gateway.save()
        else:
            messages.error(request, f"Failed to send configuration to {gateway}. Check MQTT connection.")

        return redirect('admin:core_gateway_changelist')

    # Default values for the form
    server_display = "192.168.1.92:9081"
    if gateway.app_server_ip:
        server_display = f"{gateway.app_server_ip}:{gateway.app_server_port or 9081}"

    context = {
        'gateway': gateway,
        'opts': opts,
        'server_display': server_display,
        'title': f"Configure Gateway: {gateway.estation_id}",
    }
    return render(request, 'admin/core/gateway/configure.html', context)
